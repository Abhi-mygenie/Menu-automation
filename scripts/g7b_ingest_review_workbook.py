#!/usr/bin/env python3
"""
G7B — Ingest reviewer-edited Excel workbooks back into a corrected placeholder JSON.

INPUT (read-only against canonical files):
  /app/scripts/_g7b_review/<DATASET_ID>_review.xlsx        (×5, edited by Sunil)
  /app/memory/menu-import/MENU_AI_FIRST_PASS_OUTPUT_SMOKE_v0.1.0.json
  /app/memory/menu-import/MENU_EXPECTED_OUTPUT_PLACEHOLDERS_v0.1.0.json

OUTPUT:
  /app/memory/menu-import/MENU_EXPECTED_OUTPUT_PLACEHOLDERS_v0.1.0_G7B_CORRECTED.json
  /app/memory/menu-import/G7B_REVIEW_DIFF_REPORT_v0.1.0.md

NON-GOALS
  - Does NOT mutate the canonical placeholder JSON. A new *_G7B_CORRECTED.json is
    written instead.
  - Does NOT freeze the dataset.
  - Does NOT call Gemini.
  - Does NOT sync to POS.
"""
from __future__ import annotations

import json
import time
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

MEMORY_DIR = Path("/app/memory/menu-import")
ARCHIVE_PATH = MEMORY_DIR / "MENU_AI_FIRST_PASS_OUTPUT_SMOKE_v0.1.0.json"
PLACEHOLDERS_PATH = MEMORY_DIR / "MENU_EXPECTED_OUTPUT_PLACEHOLDERS_v0.1.0.json"
OUTPUT_JSON_PATH = MEMORY_DIR / "MENU_EXPECTED_OUTPUT_PLACEHOLDERS_v0.1.0_G7B_CORRECTED.json"
DIFF_REPORT_PATH = MEMORY_DIR / "G7B_REVIEW_DIFF_REPORT_v0.1.0.md"
REVIEW_DIR = Path("/app/scripts/_g7b_review")

SMOKE_SET_IDS = [
    "MENU-v0.1.0-0007",
    "MENU-v0.1.0-0013",
    "MENU-v0.1.0-0023",
    "MENU-v0.1.0-0024",
    "MENU-v0.1.0-0025",
]

ROW_COLUMNS = [
    "row_no", "page_number", "category", "item_name", "rate", "currency",
    "confidence", "issue_status", "variant_warning", "addon_warning",
    "tax_note_warning", "raw_text", "source_bbox",
    "Action", "Corrected_item_name", "Corrected_rate", "Corrected_category",
    "Corrected_issue_status", "Reviewer_notes",
]

ALLOWED_ACTIONS = {
    "approve", "edit", "delete_hallucination",
    "add_missing", "unclear", "out_of_scope",
}

EXPECTED_ISSUE_STATUSES = {"clean", "review_required", "flagged_only_phase1"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _read_row_sheet(ws) -> List[Dict[str, Any]]:
    """Read a Page_NN sheet into list[dict]. Stops at first fully-empty row."""
    header_row = [c.value for c in ws[1]]
    header_map = {name: idx for idx, name in enumerate(header_row) if name}
    # Validate columns
    missing = [c for c in ROW_COLUMNS if c not in header_map]
    if missing:
        raise SystemExit(f"Sheet '{ws.title}' missing columns: {missing}")
    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        # Stop at fully-blank row (no row_no AND no item_name AND no Action)
        if all(v in (None, "") for v in r[: len(header_row)]):
            break
        # Also stop if we've drifted into the MENU NOTES block (handled separately)
        marker = r[0]
        if isinstance(marker, str) and marker.strip().upper().startswith("MENU NOTES"):
            break
        rec = {col: r[header_map[col]] for col in ROW_COLUMNS}
        rows.append(rec)
    return rows


def _read_menu_notes_block(ws) -> List[Dict[str, Any]]:
    """Look for a 'MENU NOTES (page-level)' block and parse its rows."""
    notes: List[Dict[str, Any]] = []
    block_start = None
    for r_idx in range(1, ws.max_row + 1):
        v = ws.cell(row=r_idx, column=1).value
        if isinstance(v, str) and "MENU NOTES" in v.upper():
            block_start = r_idx + 1
            break
    if block_start is None or block_start + 1 > ws.max_row:
        return notes
    header = [ws.cell(row=block_start, column=c).value for c in range(1, 8)]
    if not header or "note_text" not in (header[0] or ""):
        return notes
    for r_idx in range(block_start + 1, ws.max_row + 1):
        nt = ws.cell(row=r_idx, column=1).value
        if not nt:
            break
        notes.append({
            "note_text": nt,
            "note_type": ws.cell(row=r_idx, column=2).value,
            "tax_note_warning": bool(ws.cell(row=r_idx, column=3).value),
            "confidence": ws.cell(row=r_idx, column=4).value,
            "Action": (ws.cell(row=r_idx, column=5).value or "approve"),
            "Corrected_note_text": ws.cell(row=r_idx, column=6).value,
            "Reviewer_notes": ws.cell(row=r_idx, column=7).value,
        })
    return notes


def _coerce_rate(v: Any) -> Optional[str]:
    if v is None or v == "":
        return None
    try:
        f = float(v)
        if f.is_integer():
            return f"{int(f)}.00"
        return f"{f:.2f}"
    except (TypeError, ValueError):
        return str(v)


def _expected_row_from(rec: Dict[str, Any]) -> Dict[str, Any]:
    """Build the canonical expected_row dict from a (possibly corrected) review row."""
    item_name = rec.get("Corrected_item_name") or rec.get("item_name")
    rate_val = rec.get("Corrected_rate") if rec.get("Corrected_rate") not in (None, "") else rec.get("rate")
    category = rec.get("Corrected_category") or rec.get("category")
    issue_status = rec.get("Corrected_issue_status") or rec.get("issue_status")
    if issue_status not in EXPECTED_ISSUE_STATUSES:
        # remap category_inferred -> review_required for expected output (reviewer must decide)
        issue_status = "review_required" if issue_status == "category_inferred" else issue_status
    return {
        "row_index": int(rec.get("row_no") or 0),
        "category": category,
        "subcategory": None,
        "item_name": item_name,
        "rate": _coerce_rate(rate_val),
        "currency": rec.get("currency") or "INR",
        "raw_text_reference": rec.get("raw_text") or "",
        "source_grounding_required": True,
        "source_bbox_hint": None,
        "expected_issue_status": issue_status if issue_status in EXPECTED_ISSUE_STATUSES else "clean",
        "expected_warnings": [],
        "variant_flag_expected": bool(rec.get("variant_warning")),
        "addon_flag_expected": bool(rec.get("addon_warning")),
        "tax_note_flag_expected": bool(rec.get("tax_note_warning")),
        "confidence_expectation": rec.get("confidence") or "medium",
        "phase1_required_extracted": True,
        "phase2_only_detail": None,
        "notes": rec.get("Reviewer_notes") or "",
    }


def _expected_menu_note_from(note: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "note_text": note.get("Corrected_note_text") or note.get("note_text"),
        "note_type": note.get("note_type"),
        "expected_mapping": "unmapped",
        "must_be_detected_in_phase1": True,
        "notes": note.get("Reviewer_notes") or "",
    }


# ---------------------------------------------------------------------------
# Per-workbook ingest
# ---------------------------------------------------------------------------

def _ingest_workbook(path: Path) -> Tuple[List[Dict[str, Any]], Counter, Dict[str, Any]]:
    """Returns (expected_pages, action_counter, totals)."""
    wb = load_workbook(path, data_only=True)
    expected_pages: List[Dict[str, Any]] = []
    actions = Counter()
    totals = {
        "rows_seen": 0,
        "rows_in_expected": 0,
        "rows_added": 0,
        "rows_deleted": 0,
        "rows_edited": 0,
        "menu_notes_in_expected": 0,
        "unclear_remaining": 0,
        "out_of_scope": 0,
    }
    # Iterate sheets in name order so pages come out 01, 02, ...
    page_sheets = sorted(
        [s for s in wb.sheetnames if s.startswith("Page_") and s != "Page_warnings"],
        key=lambda s: int(s.split("_")[1]),
    )
    for sname in page_sheets:
        ws = wb[sname]
        page_no = int(sname.split("_")[1])
        rows = _read_row_sheet(ws)
        notes = _read_menu_notes_block(ws)
        expected_rows: List[Dict[str, Any]] = []
        for rec in rows:
            totals["rows_seen"] += 1
            action = (rec.get("Action") or "").strip().lower()
            if action and action not in ALLOWED_ACTIONS:
                raise SystemExit(
                    f"{path.name} :: sheet '{sname}' :: row_no={rec.get('row_no')} :: "
                    f"invalid Action='{action}' (allowed: {sorted(ALLOWED_ACTIONS)})"
                )
            actions[action or "(blank)"] += 1
            if action == "delete_hallucination":
                totals["rows_deleted"] += 1
                continue
            if action == "out_of_scope":
                totals["out_of_scope"] += 1
                continue
            if action == "add_missing":
                totals["rows_added"] += 1
            if action == "edit":
                totals["rows_edited"] += 1
            if action == "unclear" or action == "":
                totals["unclear_remaining"] += 1
            # approve / edit / add_missing / unclear all flow into expected
            er = _expected_row_from(rec)
            expected_rows.append(er)
            totals["rows_in_expected"] += 1

        # Renumber row_index 1..N after deletes/adds, preserve order
        for i, er in enumerate(expected_rows, start=1):
            er["row_index"] = i

        expected_menu_notes: List[Dict[str, Any]] = []
        for n in notes:
            a = (n.get("Action") or "").strip().lower()
            if a == "delete_hallucination" or a == "out_of_scope":
                continue
            expected_menu_notes.append(_expected_menu_note_from(n))
            totals["menu_notes_in_expected"] += 1

        expected_pages.append({
            "page_number": page_no,
            "expected_rows": expected_rows,
            "expected_menu_notes": expected_menu_notes,
        })

    return expected_pages, actions, totals


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    if not ARCHIVE_PATH.exists():
        raise SystemExit(f"Archive not found: {ARCHIVE_PATH}")
    placeholders = json.loads(PLACEHOLDERS_PATH.read_text())

    missing_workbooks = []
    for sid in SMOKE_SET_IDS:
        if not (REVIEW_DIR / f"{sid}_review.xlsx").exists():
            missing_workbooks.append(sid)
    if missing_workbooks:
        raise SystemExit(
            "Missing edited review workbooks for: "
            + ", ".join(missing_workbooks)
            + f". Expected each at {REVIEW_DIR}/<ID>_review.xlsx"
        )

    # Deep copy via JSON round-trip so we never mutate the loaded structure
    corrected = json.loads(json.dumps(placeholders))
    corrected["template_version"] = placeholders.get("template_version", "1.0")
    corrected["instance_purpose"] = (
        "G7B-corrected expected output for the 5 Smoke Set menus. "
        "Canonical placeholders file is NOT modified by this process."
    )
    corrected["generated_at"] = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    corrected["generated_by"] = "g7b_ingest_review_workbook.py"
    corrected["last_amendment"] = {
        "kind": "g7b_smoke_set_review",
        "reviewer": "Sunil",
        "scope": "Smoke Set only (5 of 32 menus)",
        "source_workbooks": [
            str(REVIEW_DIR / f"{sid}_review.xlsx") for sid in SMOKE_SET_IDS
        ],
    }
    corrected["last_amended_at"] = corrected["generated_at"]

    per_doc_summary: Dict[str, Any] = {}
    for sid in SMOKE_SET_IDS:
        wb_path = REVIEW_DIR / f"{sid}_review.xlsx"
        print(f"[g7b-ingest] reading {wb_path}")
        expected_pages, actions, totals = _ingest_workbook(wb_path)

        # Find target entry and populate
        entry = next(e for e in corrected["entries"]
                     if e["instance_metadata"]["dataset_id"] == sid)
        entry["expected_pages"] = expected_pages
        # Aggregate counts
        total_rows = sum(len(p["expected_rows"]) for p in expected_pages)
        total_notes = sum(len(p["expected_menu_notes"]) for p in expected_pages)
        total_cats = len({
            r["category"] for p in expected_pages for r in p["expected_rows"]
            if r.get("category")
        })
        total_variant = sum(
            1 for p in expected_pages for r in p["expected_rows"] if r.get("variant_flag_expected")
        )
        total_addon = sum(
            1 for p in expected_pages for r in p["expected_rows"] if r.get("addon_flag_expected")
        )
        total_tax = sum(
            1 for p in expected_pages for r in p["expected_rows"] if r.get("tax_note_flag_expected")
        )
        entry["expected_aggregate_counts"] = {
            "total_rows_expected": total_rows,
            "total_categories_expected": total_cats,
            "total_menu_notes_expected": total_notes,
            "total_variant_flags_expected": total_variant,
            "total_addon_flags_expected": total_addon,
            "total_tax_note_flags_expected": total_tax,
        }
        # Touch metadata: reviewer info + status (still NOT frozen)
        meta = entry["instance_metadata"]
        meta["reviewer"] = meta.get("reviewer") or "Sunil"
        meta["human_review_status"] = "G7B_COMPLETE_PENDING_G8_FREEZE"
        meta["notes"] = (meta.get("notes") or "") + " | G7B reviewer (Sunil) completed; not frozen."
        per_doc_summary[sid] = {
            "actions": dict(actions),
            "totals": totals,
            "expected_rows_total": total_rows,
            "expected_menu_notes_total": total_notes,
            "pages": len(expected_pages),
        }

    OUTPUT_JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp = OUTPUT_JSON_PATH.with_suffix(OUTPUT_JSON_PATH.suffix + ".tmp")
    tmp.write_text(json.dumps(corrected, indent=2, ensure_ascii=False))
    tmp.replace(OUTPUT_JSON_PATH)
    print(f"[g7b-ingest] wrote corrected JSON: {OUTPUT_JSON_PATH}")

    # ----- Diff report -----
    archive = json.loads(ARCHIVE_PATH.read_text())
    docs_by_id = {d["import_id"].replace("smoke-", ""): d for d in archive["documents"]}
    lines: List[str] = []
    lines.append("# G7B Review Diff Report — v0.1.0")
    lines.append("")
    lines.append(f"- Generated: {corrected['generated_at']}")
    lines.append("- Reviewer: Sunil")
    lines.append("- Scope: Smoke Set only (5 of 32 menus). Other 27 menus not processed.")
    lines.append(f"- Canonical placeholder file untouched: `{PLACEHOLDERS_PATH}`")
    lines.append(f"- Output (separate file): `{OUTPUT_JSON_PATH}`")
    lines.append("- Dataset NOT frozen. POS NOT synced. No Gemini call.")
    lines.append("")
    lines.append("## Per-document summary")
    lines.append("")
    lines.append("| Dataset ID | Source file | AI rows | Expected rows | Δ | Approved | Edited | Deleted | Added | Unclear | Out-of-scope |")
    lines.append("|---|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|")
    grand = Counter()
    for sid in SMOKE_SET_IDS:
        d = docs_by_id[sid]
        ai_rows = sum(len(p["rows"]) for p in d["pages"])
        s = per_doc_summary[sid]
        a = s["actions"]
        exp = s["expected_rows_total"]
        delta = exp - ai_rows
        approved = a.get("approve", 0)
        edited = a.get("edit", 0)
        deleted = a.get("delete_hallucination", 0)
        added = a.get("add_missing", 0)
        unclear = a.get("unclear", 0) + a.get("(blank)", 0)
        oos = a.get("out_of_scope", 0)
        source = next(m["source_file"] for m in archive["per_document_metadata"]
                      if m["dataset_id"] == sid)
        lines.append(
            f"| {sid} | {source} | {ai_rows} | {exp} | {delta:+d} | "
            f"{approved} | {edited} | {deleted} | {added} | {unclear} | {oos} |"
        )
        grand["approve"] += approved
        grand["edit"] += edited
        grand["delete_hallucination"] += deleted
        grand["add_missing"] += added
        grand["unclear"] += unclear
        grand["out_of_scope"] += oos
        grand["ai_rows"] += ai_rows
        grand["expected_rows"] += exp
    lines.append("")
    lines.append("**Totals:** "
                 f"ai_rows={grand['ai_rows']} | expected_rows={grand['expected_rows']} | "
                 f"approve={grand['approve']} | edit={grand['edit']} | "
                 f"delete={grand['delete_hallucination']} | add={grand['add_missing']} | "
                 f"unclear={grand['unclear']} | out_of_scope={grand['out_of_scope']}")
    lines.append("")
    lines.append("## Action interpretation")
    lines.append("")
    lines.append("- **approve** — AI row kept verbatim. Counts toward ground truth.")
    lines.append("- **edit** — AI row kept with `Corrected_*` overrides applied. Counts toward ground truth.")
    lines.append("- **delete_hallucination** — AI invented this row; it does NOT appear in ground truth.")
    lines.append("- **add_missing** — Reviewer added a row the AI missed. Pre-populated as a new entry; reviewer must fill `Corrected_*`. Counts toward ground truth.")
    lines.append("- **unclear** — Reviewer left undecided. Kept in ground truth as `review_required`. Should be reviewed again before G8 freeze.")
    lines.append("- **out_of_scope** — Row excluded from this dataset version (not menu content, ads, etc.). Does not count.")
    lines.append("")
    lines.append("## Next gates")
    lines.append("")
    lines.append("1. **Re-review `unclear` rows** if any remain (see column above).")
    lines.append("2. **G7C** — AI first pass on Learning Memory Set (NOT yet started). Gemini call gated.")
    lines.append("3. **G7D** — Human review for Learning Memory Set.")
    lines.append("4. **G7E/G7F** — Eval Hold-out (once-only AI pass + human review).")
    lines.append("5. **G8** — Dataset freeze → v0.1.0 immutable. Will copy `expected_pages` from this corrected file into the canonical placeholders + set `frozen_at`.")
    lines.append("")
    DIFF_REPORT_PATH.write_text("\n".join(lines))
    print(f"[g7b-ingest] wrote diff report: {DIFF_REPORT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
