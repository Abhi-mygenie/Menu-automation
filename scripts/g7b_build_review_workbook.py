#!/usr/bin/env python3
"""
G7B — Build per-document Excel review workbooks for the 5 Smoke Set menus.

INPUT (read-only):
  /app/memory/menu-import/MENU_AI_FIRST_PASS_OUTPUT_SMOKE_v0.1.0.json
  /app/scripts/_g7a_staging/extraction_log/MENU-v0.1.0-XXXX.json
  /app/memory/menu-import/MENU_EXPECTED_OUTPUT_PLACEHOLDERS_v0.1.0.json (for sha + source paths)

OUTPUT:
  /app/scripts/_g7b_review/<DATASET_ID>_review.xlsx        (×5)
  /app/scripts/_g7b_review/_build_summary.json

NON-GOALS
  - Does NOT call Gemini.
  - Does NOT mutate the AI archive or the canonical placeholder JSON.
  - Does NOT process the other 27 menus.
  - Does NOT freeze the dataset.
  - Does NOT sync to POS.
  - Does NOT implement the production review UI.
"""
from __future__ import annotations

import json
import time
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

MEMORY_DIR = Path("/app/memory/menu-import")
ARCHIVE_PATH = MEMORY_DIR / "MENU_AI_FIRST_PASS_OUTPUT_SMOKE_v0.1.0.json"
PLACEHOLDERS_PATH = MEMORY_DIR / "MENU_EXPECTED_OUTPUT_PLACEHOLDERS_v0.1.0.json"
LOG_DIR = Path("/app/scripts/_g7a_staging/extraction_log")
OUT_DIR = Path("/app/scripts/_g7b_review")

SMOKE_SET_IDS = [
    "MENU-v0.1.0-0007",
    "MENU-v0.1.0-0013",
    "MENU-v0.1.0-0023",
    "MENU-v0.1.0-0024",
    "MENU-v0.1.0-0025",
]

ACTION_VALUES = [
    "approve",
    "edit",
    "delete_hallucination",
    "add_missing",
    "unclear",
    "out_of_scope",
]

# Column order on the per-page row sheets
ROW_COLUMNS = [
    "row_no",
    "page_number",
    "category",
    "item_name",
    "rate",
    "currency",
    "confidence",
    "issue_status",
    "variant_warning",
    "addon_warning",
    "tax_note_warning",
    "raw_text",
    "source_bbox",
    "Action",
    "Corrected_item_name",
    "Corrected_rate",
    "Corrected_category",
    "Corrected_issue_status",
    "Reviewer_notes",
]

# Cell fills
FILL_HEADER = PatternFill("solid", fgColor="1F2937")     # dark slate
FILL_REVIEW = PatternFill("solid", fgColor="FEE2E2")     # red-100
FILL_FLAGGED = PatternFill("solid", fgColor="FEF3C7")    # amber-100
FILL_CLEAN = PatternFill("solid", fgColor="ECFDF5")      # emerald-50
FILL_CAT_INFER = PatternFill("solid", fgColor="EDE9FE")  # violet-100
FONT_HEADER = Font(color="FFFFFF", bold=True)
FONT_BOLD = Font(bold=True)

# Pre-fill rules
def _prefill_action(row: Dict[str, Any]) -> str:
    if row["issue_status"] == "clean" and row["confidence"] in ("high", "medium"):
        return "approve"
    if row["issue_status"] in ("review_required", "flagged_only_phase1"):
        return "unclear"
    if row["issue_status"] == "category_inferred":
        return "unclear"
    if row["confidence"] == "low":
        return "unclear"
    return "approve"


# ---------------------------------------------------------------------------
# Workbook builders
# ---------------------------------------------------------------------------

def _autosize(ws, max_widths: Dict[str, int]) -> None:
    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        letter = get_column_letter(col_idx)
        header = ws.cell(row=1, column=col_idx).value or ""
        cap = max_widths.get(header, 40)
        longest = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=len(str(header)),
        )
        ws.column_dimensions[letter].width = min(max(12, longest + 2), cap)


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


def _bbox_str(b: Any) -> str:
    if not b:
        return ""
    if isinstance(b, dict):
        return f"x={b.get('x')}, y={b.get('y')}, w={b.get('w')}, h={b.get('h')}"
    return str(b)


def _add_action_dropdown(ws, action_col_letter: str, start_row: int, end_row: int) -> None:
    formula = '"' + ",".join(ACTION_VALUES) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    dv.error = f"Action must be one of: {', '.join(ACTION_VALUES)}"
    dv.errorTitle = "Invalid Action"
    dv.prompt = f"Pick: {' | '.join(ACTION_VALUES)}"
    dv.promptTitle = "G7B Review Action"
    ws.add_data_validation(dv)
    dv.add(f"{action_col_letter}{start_row}:{action_col_letter}{end_row}")


def _build_summary_sheet(ws, doc: Dict[str, Any], doc_meta: Dict[str, Any],
                         ext_log: Dict[str, Any] | None, placeholder_meta: Dict[str, Any]) -> None:
    ws.title = "Summary"
    rows: List[List[Any]] = []
    rows.append(["Field", "Value"])
    rows.append(["dataset_id", placeholder_meta["dataset_id"]])
    rows.append(["source_file", placeholder_meta["source_file"]])
    rows.append(["dataset_version", placeholder_meta["dataset_version"]])
    rows.append(["sha256", placeholder_meta["sha256"]])
    rows.append(["local_storage_path", placeholder_meta.get("local_storage_path", "")])
    rows.append(["page_count", doc_meta["page_count"]])
    rows.append(["envelopes_sent_to_gemini", doc_meta["envelopes_sent"]])
    rows.append(["chunked", "yes" if doc_meta["chunked"] else "no"])
    rows.append(["ocr_low_confidence_flag", "yes" if doc_meta.get("ocr_low_confidence") else "no"])
    rows.append(["", ""])

    # Issue status / confidence rollups
    from collections import Counter
    is_counts = Counter()
    cf_counts = Counter()
    total_rows = 0
    total_menu_notes = 0
    for p in doc["pages"]:
        for r in p["rows"]:
            is_counts[r["issue_status"]] += 1
            cf_counts[r["confidence"]] += 1
            total_rows += 1
        total_menu_notes += len(p.get("menu_notes") or [])
    rows.append(["AI extraction_summary", ""])
    for k, v in doc.get("extraction_summary", {}).items():
        rows.append([f"  {k}", v])
    rows.append(["", ""])
    rows.append(["row issue_status counts", ""])
    for k in ("clean", "review_required", "flagged_only_phase1", "category_inferred"):
        rows.append([f"  {k}", is_counts.get(k, 0)])
    rows.append(["row confidence counts", ""])
    for k in ("high", "medium", "low"):
        rows.append([f"  {k}", cf_counts.get(k, 0)])
    rows.append(["", ""])
    rows.append(["top-level warnings (document)", ", ".join(doc.get("warnings") or []) or "(none)"])
    if doc_meta.get("warnings_injected"):
        rows.append(["warnings injected by G7A orchestrator",
                     ", ".join(doc_meta["warnings_injected"])])

    # Per-page extraction method (from G7A log)
    if ext_log and "pages" in ext_log:
        rows.append(["", ""])
        rows.append(["per-page extraction (G7A pipeline)", ""])
        for p in ext_log["pages"]:
            rows.append([
                f"  page {p['page_number']:02d}",
                f"method={p.get('method')} chars={p.get('char_count')} letter_ratio={p.get('letter_ratio')}",
            ])

    # Write
    for r in rows:
        ws.append(r)
    _style_header(ws, 2)
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=1)
        if cell.value and not str(cell.value).startswith("  ") and cell.value != "":
            cell.font = FONT_BOLD
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 80


def _build_page_sheet(wb: Workbook, page_no: int, page: Dict[str, Any]) -> None:
    title = f"Page_{page_no:02d}"
    ws = wb.create_sheet(title=title)

    # Header
    ws.append(ROW_COLUMNS)
    _style_header(ws, len(ROW_COLUMNS))

    # Rows
    for r in page["rows"]:
        action = _prefill_action(r)
        row_vals = [
            r["row_no"],
            page_no,
            r["category"],
            r["item_name"],
            r["rate"],
            r["currency"],
            r["confidence"],
            r["issue_status"],
            r["variant_warning"],
            r["addon_warning"],
            r["tax_note_warning"],
            r["raw_text"],
            _bbox_str(r.get("source_bbox")),
            action,
            "",  # Corrected_item_name
            "",  # Corrected_rate
            "",  # Corrected_category
            "",  # Corrected_issue_status
            "",  # Reviewer_notes
        ]
        ws.append(row_vals)

    # Row-level fills (issue_status driven)
    issue_col_idx = ROW_COLUMNS.index("issue_status") + 1
    for r_idx in range(2, ws.max_row + 1):
        status = ws.cell(row=r_idx, column=issue_col_idx).value
        fill = None
        if status == "review_required":
            fill = FILL_REVIEW
        elif status == "flagged_only_phase1":
            fill = FILL_FLAGGED
        elif status == "category_inferred":
            fill = FILL_CAT_INFER
        elif status == "clean":
            fill = FILL_CLEAN
        if fill:
            ws.cell(row=r_idx, column=issue_col_idx).fill = fill

    # Filters + frozen header
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ROW_COLUMNS))}{max(ws.max_row, 2)}"

    # Drop-down for Action column
    action_col_idx = ROW_COLUMNS.index("Action") + 1
    action_col_letter = get_column_letter(action_col_idx)
    if ws.max_row >= 2:
        _add_action_dropdown(ws, action_col_letter, 2, ws.max_row + 200)
        # also drop-down for Corrected_issue_status
        cis_idx = ROW_COLUMNS.index("Corrected_issue_status") + 1
        cis_letter = get_column_letter(cis_idx)
        dv2 = DataValidation(
            type="list",
            formula1='"clean,review_required,flagged_only_phase1,category_inferred"',
            allow_blank=True,
        )
        dv2.prompt = "Pick a corrected issue_status (leave blank to keep original)"
        ws.add_data_validation(dv2)
        dv2.add(f"{cis_letter}2:{cis_letter}{ws.max_row + 200}")

    # Wrap raw_text + word-wrap header
    raw_col_idx = ROW_COLUMNS.index("raw_text") + 1
    raw_col_letter = get_column_letter(raw_col_idx)
    ws.column_dimensions[raw_col_letter].width = 60
    for r_idx in range(2, ws.max_row + 1):
        ws.cell(row=r_idx, column=raw_col_idx).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r_idx, column=ROW_COLUMNS.index("category") + 1).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.cell(row=r_idx, column=ROW_COLUMNS.index("item_name") + 1).alignment = Alignment(
            wrap_text=True, vertical="top"
        )

    # Conditional formatting on Action column (visual cue once filled)
    action_letter = get_column_letter(action_col_idx)
    last_action_row = ws.max_row + 200
    ws.conditional_formatting.add(
        f"{action_letter}2:{action_letter}{last_action_row}",
        CellIsRule(operator="equal", formula=['"approve"'],
                   fill=PatternFill("solid", fgColor="D1FAE5")),  # emerald-100
    )
    ws.conditional_formatting.add(
        f"{action_letter}2:{action_letter}{last_action_row}",
        CellIsRule(operator="equal", formula=['"edit"'],
                   fill=PatternFill("solid", fgColor="DBEAFE")),  # blue-100
    )
    ws.conditional_formatting.add(
        f"{action_letter}2:{action_letter}{last_action_row}",
        CellIsRule(operator="equal", formula=['"delete_hallucination"'],
                   fill=PatternFill("solid", fgColor="FECACA")),  # red-200
    )
    ws.conditional_formatting.add(
        f"{action_letter}2:{action_letter}{last_action_row}",
        CellIsRule(operator="equal", formula=['"add_missing"'],
                   fill=PatternFill("solid", fgColor="FDE68A")),  # amber-200
    )
    ws.conditional_formatting.add(
        f"{action_letter}2:{action_letter}{last_action_row}",
        CellIsRule(operator="equal", formula=['"unclear"'],
                   fill=PatternFill("solid", fgColor="E5E7EB")),  # gray-200
    )
    ws.conditional_formatting.add(
        f"{action_letter}2:{action_letter}{last_action_row}",
        CellIsRule(operator="equal", formula=['"out_of_scope"'],
                   fill=PatternFill("solid", fgColor="F3E8FF")),  # violet-100
    )

    # Column widths
    _autosize(ws, max_widths={
        "raw_text": 60,
        "category": 28,
        "item_name": 36,
        "Reviewer_notes": 40,
        "Corrected_item_name": 36,
        "Corrected_category": 28,
        "source_bbox": 26,
    })

    # Add menu_notes block at the bottom if any
    notes = page.get("menu_notes") or []
    if notes:
        ws.append([])
        ws.append(["MENU NOTES (page-level)"])
        ws.cell(row=ws.max_row, column=1).font = FONT_BOLD
        ws.append(["note_text", "note_type", "tax_note_warning", "confidence",
                   "Action", "Corrected_note_text", "Reviewer_notes"])
        _style_header(ws, 7)  # no-op for older sheet, but harmless
        notes_header_row = ws.max_row
        for n in notes:
            ws.append([
                n["note_text"], n["note_type"], n["tax_note_warning"], n["confidence"],
                "approve", "", "",
            ])
        # Action drop-down on the notes block
        note_action_col = 5
        note_action_letter = get_column_letter(note_action_col)
        dv3 = DataValidation(
            type="list",
            formula1='"approve,edit,delete_hallucination,out_of_scope,unclear"',
            allow_blank=True,
        )
        ws.add_data_validation(dv3)
        dv3.add(f"{note_action_letter}{notes_header_row + 1}:"
                f"{note_action_letter}{notes_header_row + 1 + len(notes) + 50}")


def _build_page_warnings_sheet(wb: Workbook, doc: Dict[str, Any]) -> None:
    ws = wb.create_sheet(title="Page_warnings")
    ws.append(["page_number", "warnings", "row_count", "menu_note_count"])
    _style_header(ws, 4)
    for p in doc["pages"]:
        ws.append([
            p["page_number"],
            ", ".join(p.get("warnings") or []) or "(none)",
            len(p["rows"]),
            len(p.get("menu_notes") or []),
        ])
    ws.auto_filter.ref = f"A1:D{max(ws.max_row, 2)}"
    _autosize(ws, max_widths={"warnings": 80})


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    if not ARCHIVE_PATH.exists():
        raise SystemExit(f"Archive not found: {ARCHIVE_PATH}")
    archive = json.loads(ARCHIVE_PATH.read_text())
    placeholders = json.loads(PLACEHOLDERS_PATH.read_text())
    placeholder_index = {
        e["instance_metadata"]["dataset_id"]: e["instance_metadata"]
        for e in placeholders["entries"]
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    summary = {
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "archive_source": str(ARCHIVE_PATH),
        "workbooks": [],
        "scope": "G7B Smoke Set review only (5 of 32 menus)",
        "non_goals": [
            "No Gemini call",
            "No dataset freeze",
            "No POS sync",
            "No mutation of canonical placeholder JSON",
            "No processing of the remaining 27 menus",
        ],
    }

    docs_by_id = {d["import_id"].replace("smoke-", ""): d for d in archive["documents"]}
    meta_by_id = {m["dataset_id"]: m for m in archive["per_document_metadata"]}

    for sid in SMOKE_SET_IDS:
        if sid not in docs_by_id:
            raise SystemExit(f"Document {sid} missing from archive.")
        if sid not in placeholder_index:
            raise SystemExit(f"Placeholder entry missing for {sid}.")
        doc = docs_by_id[sid]
        doc_meta = meta_by_id[sid]
        ext_log = None
        log_path = LOG_DIR / f"{sid}.json"
        if log_path.exists():
            ext_log = json.loads(log_path.read_text())
        placeholder_meta = placeholder_index[sid]

        wb = Workbook()
        # default sheet -> Summary
        _build_summary_sheet(wb.active, doc, doc_meta, ext_log, placeholder_meta)
        for page in doc["pages"]:
            _build_page_sheet(wb, page["page_number"], page)
        _build_page_warnings_sheet(wb, doc)

        out_path = OUT_DIR / f"{sid}_review.xlsx"
        wb.save(out_path)
        summary["workbooks"].append({
            "dataset_id": sid,
            "source_file": placeholder_meta["source_file"],
            "path": str(out_path),
            "pages": len(doc["pages"]),
            "rows": sum(len(p["rows"]) for p in doc["pages"]),
            "menu_notes": sum(len(p.get("menu_notes") or []) for p in doc["pages"]),
            "ocr_low_confidence": doc_meta.get("ocr_low_confidence", False),
        })
        print(f"[g7b] wrote {out_path}  pages={len(doc['pages'])}  "
              f"rows={sum(len(p['rows']) for p in doc['pages'])}")

    (OUT_DIR / "_build_summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False))
    print(f"[g7b] build summary: {OUT_DIR / '_build_summary.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
